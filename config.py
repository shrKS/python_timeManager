import openpyxl
import os
import timerDB

from sql_python import DB_NAME
EXCEL_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), "test_book.xlsx")
DATABASE = "timeManageData"
TITLE = "test"
WINDOW_WIDTH=3000
WINDOW_HEGHT=500
COLUMNS = {}
WB = openpyxl.load_workbook(EXCEL_PATH)
DB_SHEET = WB[DATABASE]


#for postgresql info
DB_HOST = 'dbmanager'
DB_NAME = 'test_db'
DB_TABLE = 'Timer'
TIMERDB = timerDB.DataBase(DB_HOST, DB_NAME, DB_TABLE)

col = 1
while DB_SHEET.cell(row=1, column=col).value is not None:
    COLUMNS[DB_SHEET.cell(row=1, column=col).value] = col
    col+=1
print(COLUMNS)