import openpyxl
from datetime import datetime
import config

class stopWatch:
    def __init__(self, wb, sheet):
        self.wb = wb
        self.sheet = sheet
        self.columns = config.COLUMNS
        self.setTitle("test")
    
    def setTitle(self, title):
        self.title = title

    def start(self):
        self.s_time = datetime.now()

    def stop(self):
        self.e_time = datetime.now()
        self.calcDiff()
        self.record()
        self.resetTime()

    def calcDiff(self):
        self.diff_time = self.e_time - self.s_time
        print(self.diff_time)

    def resetTime(self):
        self.s_time = None
        self.e_time = None
        self.diff_time = None

    def record(self):
        empty_row = self.searchEmptyRow()

        #record to excel
        self.sheet.cell(row=empty_row, column=self.columns['index'], value=empty_row-1)
        self.sheet.cell(row=empty_row, column=self.columns['題目'], value=self.title)
        self.sheet.cell(row=empty_row, column=self.columns['開始時刻'], value=self.s_time)
        self.sheet.cell(row=empty_row, column=self.columns['終了時刻'], value=self.e_time)
        self.sheet.cell(row=empty_row, column=self.columns['時間'], value=self.diff_time)

        #save
        self.wb.save(config.EXCEL_PATH)


    def searchEmptyRow(self):
        #excel index is start at 1
        row=1
        while not self.sheet.cell(row,self.columns['index']).value is None:
            row+=1
        return row