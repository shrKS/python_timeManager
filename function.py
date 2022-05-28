import pandas as pd
import openpyxl
from datetime import datetime as datetime
import stopWatch

def testCommand(filename):
    print(filename)
    input_book = pd.ExcelFile(filename)
    input_sheet_name = input_book.sheet_names
    print(input_sheet_name)
    
def searchEmptyRow(sheet, indexCol):
    #excel index is start at 1
    row=1
    while not sheet.cell(row,indexCol).value is None:
        row+=1
    return row

# def inputTime(sheet, row, indexCol):
#     now_time = datetime.now()
#     sheet.cell(row, indexCol, value=row-1)
#     sheet.cell(row, indexCol+1, value=now_time)
#     hours = "{}:{}".format(now_time.hour, now_time.minute)
#     sheet.cell(row, indexCol+2, value=hours)

# def inputData(filename):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['timeManageData']
#     print(sheet.cell(1,1).value)
#     row = searchEmptyRow(sheet, indexCol=1)
#     print(row)
#     inputTime(sheet, row, indexCol=1)
#     wb.save(filename)

# def outputTime(sheet, row, indexCol):
#     now_time = datetime.now()
#     sheet.cell(row, indexCol+3, value=now_time)
#     time_diff = sheet.cell(row, indexCol+3).value - sheet.cell(row, indexCol+1).value
#     sheet.cell(row, indexCol+4, value=time_diff)
    


# def outputData(filename):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['timeManageData']
#     row = searchEmptyRow(sheet, indexCol=1) -1
#     outputTime(sheet, row, indexCol=1)
#     wb.save(filename)



def timeStart(watch):
    watch.start()

def timeStop(watch):
     watch.stop()

    

if __name__ == '__main__':
    print("package")


#メモ

#タイトルの追加や表示の設定(クラス?)
#groupbyしてグラフ表示(棒グラフ，チャート)
#休憩時間の開始と終了の追加
#6つの枠で画面を満たす
